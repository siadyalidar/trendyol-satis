"""
cost_import.py
---------------
"🧮 Ürün Maliyet" sekmesindeki ürün maliyetlerini SQLite'a aktarır.

Beklenen sütun sırası (Excel'inizdeki başlıklarla eşleşir):
    A: SKU
    B: Ürün Adı
    C: Satış Fiyatı (KDV dahil)
    D: Alış Fiyatı (KDV dahil)
    E: Satış (KDV hariç)
    F: Alış (KDV hariç)
    ... (G'den sonrası -TY/HB Hesaba Yatan, Brüt Kâr, Marj%, ROI, Durum- bilgi
    amaçlı, DB'ye aktarılmıyor; bunlar profit_engine.py'de yeniden hesaplanacak
    çünkü artık gerçek Finans API verisiyle çalışıyoruz, Excel'deki sabit
    formülle değil.)

Kullanım:
    python cost_import.py "SoftHydra.xlsx"
    python cost_import.py "SoftHydra.xlsx" --sheet "🧮 Ürün Maliyet"
"""

import argparse
import sys

import openpyxl

from database import init_db, upsert_product_costs

DEFAULT_SHEET = "🧮 Ürün Maliyet"


def import_product_costs(excel_path, sheet_name=DEFAULT_SHEET):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"'{sheet_name}' sekmesi bulunamadı. Mevcut sekmeler: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    rows = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = row[0]
        if not sku:
            continue
        try:
            rows.append({
                "sku": str(sku).strip(),
                "product_name": row[1],
                "sale_price_incl_vat": float(row[2]) if row[2] is not None else None,
                "cost_incl_vat": float(row[3]) if row[3] is not None else None,
                "sale_price_excl_vat": float(row[4]) if row[4] is not None else None,
                "cost_excl_vat": float(row[5]) if row[5] is not None else None,
            })
        except (TypeError, ValueError):
            skipped += 1
            continue

    init_db()
    upsert_product_costs(rows)
    return len(rows), skipped


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ürün maliyetlerini Excel'den DB'ye aktar")
    parser.add_argument("excel_path", help="Excel dosyasının yolu")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Ürün maliyet sekmesinin adı")
    args = parser.parse_args()

    try:
        imported, skipped = import_product_costs(args.excel_path, args.sheet)
    except Exception as e:
        print(f"HATA: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"{imported} ürün maliyeti aktarıldı." + (f" ({skipped} satır atlandı — eksik/hatalı veri)" if skipped else ""))
